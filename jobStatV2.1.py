#! python3
### this program gets the full ARN from appt scheduler and then gets the job stat from WFA
### by pulling the data from a user specified excel document and saving the results to a new excel doc.

import openpyxl
import os
import pyautogui
import pyperclip
import pyscreeze
import selenium
import sys
import textmyself
import time
import tkinter
from openpyxl import load_workbook
from tkinter import *
from tkinter.filedialog import askopenfilename
from tkinter.filedialog import asksaveasfile
from tkinter.messagebox import *
pyautogui.FAILSAFE = True

textFile = ''
jStatFileName = ''
name = ''
center = ''
eRow = 3
maxRow = ''
eRow2 = 3
resolution = ''
wfaIcon = ''
ieTaskBar= ''
ieWindow= ''
apptManager= ''
waitLoad11= ''
waitLoad22= ''
searchModDel= ''
arnBox= ''
searchButton= ''
arnErrOkay= ''
checkARN1= ''
checkARN2= ''
canApp= ''
appNF= ''
waitLoad3= ''
waitLoad4= ''
highlightOrNum= ''
checkApp1= ''
checkApp2= ''
checkWFAclear1= ''
checkWFAclear2= ''
wfaCenter= ''
jStatDown= ''
jStatUp= ''
centerDown= ''
centerUp= ''
ISWR= ''

def openReportFile():
    global textFile
    global jStatFileName
    global name
    try:    
        name = askopenfilename()
    except FileNotFoundError:
        print('User cancelled during file selection.')
    jStatFileName = name[:49] + '_jStat.xlsx' 

skipOrder = False
orderCancelled = False
arnError = False
aNF = False


def setResolution():
    global wfaIcon
    global ieTaskBar
    global ieWindow
    global apptManager
    global waitLoad11
    global waitLoad22
    global searchModDel
    global arnBox
    global searchButton
    global arnErrOkay
    global checkARN1
    global checkARN2
    global canApp
    global appNF
    global waitLoad3
    global waitLoad4
    global highlightOrNum
    global checkApp1
    global checkApp2
    global checkWFAclear1
    global checkWFAclear2
    global wfaCenter
    global jStatDown
    global jStatUp
    global centerDown
    global centerUp
    global ISWR
    if resolution == 1680:
        wfaIcon = (150, 1028)
        ieTaskBar = (92, 1027)
        ieWindow = (105, 952)
        apptManager = (84, 252)
        waitLoad11 = [438, 313, 0, 0, 0]
        waitLoad22 = [438, 313, 255, 255, 255]
        searchModDel = (102, 292)
        arnBox = (741, 305)
        searchButton = (634, 433)
        arnErrOkay = (952, 589)
        checkARN1 = (738, 527, 231, 166, 0)
        checkARN2 = (739, 527, 231, 164, 0)
        canApp = (558, 275, 255, 255, 255)
        appNF = (441, 282, 244, 0, 0)
        waitLoad3 = [441, 490, 0, 0, 0]
        waitLoad4 = [441, 490, 255, 255, 255]
        highlightOrNum = (673, 483)
        checkApp1 = (325, 215, 255, 204, 0)
        checkApp2 = (479, 311, 0, 102, 212)
        checkWFAclear1 = (160, 213, 0, 255, 0)
        checkWFAclear2 =(153, 209, 0, 255, 0)
        wfaCenter = (72, 176)
        jStatDown = (1112, 195)
        jStatUp = (1174, 217)
        centerDown = (20, 159)
        centerUp = (148, 186)
        ISWR = (253, 101)
    
        
def click(x,y):
    pyautogui.click(x,y)


def pixelMatch(x, y, r, g, b):
    pyscreeze.pixel(x, y) == (r, g, b)


def copy():
    pyautogui.hotkey('ctrl', 'c')

    
def type2(x):
    pyautogui.typewrite(x)


def sleep(x):
    time.sleep(x)


def wfaTaskbar():
    click(wfaIcon[0],wfaIcon[1])     


def wait3Seconds():
    print('3...')
    time.sleep(1)
    print('2...')
    time.sleep(1)
    print('1...')
    time.sleep(1)


def ieWindowClick():
    click(ieTaskBar[0],ieTaskBar[1]) # ie taskbar
    click(ieWindow[0],ieWindow[1]) # ie window (if multiple tabs - on first) 


def ieRefresh():
    pyautogui.press('f5')
    


def apptManSetUp():
    ieWindowClick()
    pyautogui.hotkey('ctrl','t')
    sleep(.5)
    ieWindowClick()
    ieRefresh()
    sleep(3)
    click(apptManager[0],apptManager[1])  # appt manager
    sleep(1)


def waitLoad1():
    for i in range(3):
        im = pyscreeze.screenshot()
        if im.getpixel((waitLoad11[0], waitLoad11[1])) == (waitLoad11[2], waitLoad11[3], waitLoad11[4]):
            break
        elif im.getpixel((waitLoad22[0],waitLoad22[1])) == (waitLoad22[2], waitLoad22[3], waitLoad22[4]):
            sleep(.2)
            if im.getpixel((waitLoad11[0], waitLoad11[1])) == (waitLoad11[2], waitLoad11[3], waitLoad11[4]):
                ieRefresh()
                getSearchBox()
            elif im.getpixel((waitLoad22[0],waitLoad22[1])) == (waitLoad22[2], waitLoad22[3], waitLoad22[4]):
                break
    
    
def getSearchBox():
    click(searchModDel[0],searchModDel[1]) #search/mod/del
    sleep(.9)
    waitLoad1()
    click(arnBox[0],arnBox[1]) #arn box


def clickSearchButton():
    click(searchButton[0],searchButton[1]) #search button


def arnErrorOkayBox():
    click(arnBox[0],arnBox[1])
    global skipOrder
    skipOrder = True


def checkArnErrorOrCancel():
    sleep(.15)
    global arnError
    global orderCancelled
    im = pyscreeze.screenshot()
    if im.getpixel((checkARN1[0],checkARN1[1])) == (checkARN1[2],checkARN1[3],checkARN1[4]) or im.getpixel((checkARN2[0],checkARN2[1])) == (checkARN2[2],checkARN2[3],checkARN2[4]):
        arnErrorOkayBox()
        arnError = True
    if im.getpixel((canApp[0], canApp[1])) != (canApp[2], canApp[3], canApp[4]):
        orderCancelled = True


def appNotFound():
    global aNF
    im = pyscreeze.screenshot()
    if im.getpixel((appNF[0],appNF[1])) == (appNF[2],appNF[3],appNF[4]):
        aNF = True
    
   
def waitLoad2():
    global skipOrder
    for i in range(3):
        im = pyscreeze.screenshot()
        if im.getpixel((waitLoad3[0], waitLoad3[1])) == (waitLoad3[2], waitLoad3[3], waitLoad3[4]):
            skipOrder = False
            break
        if im.getpixel((waitLoad4[0], waitLoad4[1])) == (waitLoad4[2], waitLoad4[3], waitLoad4[4]):
            sleep(.11)   
            skipOrder = True
            

def exceptPermissionError():
    closeWorkbook = ''
    while closeWorkbook == '':
        closeWorkbook = pyautogui.confirm(text='An instance of the file you are trying to save is already open.', title='Permission Error', buttons=['Try again', 'Exit'])
        if closeWorkbook == 'Try again':
            sleep(.15)
            ieWindowClick()
        if closeWorkbook == 'Exit':
            closeWorkbook = pyautogui.confirm(text='Are you sure you want to exit?.', title='Permission Error', buttons=['Go back', 'Exit'])
            if closeWorkbook == 'Go back':
                closeWorkbook = ''
            if closeWorkbook == 'Exit':
                sys.exit()


def exceptPermissionError2():
    closeWorkbook = ''
    while closeWorkbook == '':
        closeWorkbook = pyautogui.confirm(text='An instance of the file you are trying to save is already open.', title='Permission Error', buttons=['Try again', 'Exit'])
        if closeWorkbook == 'Try again':
            sleep(.15)
            wfaTaskbar()
        if closeWorkbook == 'Exit':
            closeWorkbook = pyautogui.confirm(text='Are you sure you want to exit?.', title='Permission Error', buttons=['Go back', 'Exit'])
            if closeWorkbook == 'Go back':
                closeWorkbook = ''
            if closeWorkbook == 'Exit':
                sys.exit()

def highlightOrderNumber():
    pyautogui.doubleClick(highlightOrNum[0],highlightOrNum[1]) # order number


def checkApptSched():
    im = pyscreeze.screenshot()
    checkApptPrompt = 'placeholder'
    if im.getpixel((checkApp1[0],checkApp1[1])) != (checkApp1[2],checkApp1[3],checkApp1[4]) or im.getpixel((checkApp2[0],checkApp2[1])) != (checkApp1[2],checkApp1[3],checkApp1[4]):
        checkApptPrompt = ''
    while checkApptPrompt == '':
        checkApptPrompt = pyautogui.confirm(text="Appointment scheduler is not properly configured. Please make sure IE is the leftmost \
icon in the taskbar, and that appointment scheduler is the leftmost tab within IE. Click 'Check Again' once you have done so. If this \
does not fix the issue, your program may need to be corrected.", title="Fix Appointment Scheduler", buttons= ['Check Again', 'Exit'])
    if checkApptPrompt == 'Check Again':
        if im.getpixel((checkApp1[0],checkApp1[1])) != (checkApp1[2],checkApp1[3],checkApp1[4]) or im.getpixel((checkApp2[0],checkApp2[1])) != (checkApp2[2],checkApp2[3],checkApp2[4]):
            checkApptPrompt = ''
    if checkApptPrompt == 'Exit':
        checkApptPrompt == pyautogui.confirm(text="Are you sure you want to exit the program?", title="Fix Appointment Scheduler", \
                                             buttons=['Yes, Exit', 'No, go back'])
    if checkApptPrompt == 'Yes, Exit':
        sys.exit()
    if checkApptPrompt == 'No, go back':
        checkApptPrompt == ''
        

def excelLoop1():
    wb = load_workbook(filename = name)
    wb.get_sheet_names()
    ws = wb.get_sheet_by_name('Report Data')
    global eRow
    global maxRow
    maxRow = ws.max_row-2
    checkApptSched()
    for i in range(ws.max_row-2):
        global skipOrder
        global orderCancelled
        global arnError
        global aNF
        skipOrder = False
        orderCancelled = False
        arnError = False
        aNF = False
        getSearchBox()
        try:
            type2(ws.cell(row=eRow, column =7).value)
        except TypeError: 
            textmyself.textmyself('Blank ARN - Break')
            break
        clickSearchButton()
        sleep(.2)
        checkArnErrorOrCancel()
        appNotFound()
        if arnError == True:
            ws.cell(row=eRow, column=8).value = 'ARN Error'
            eRow += 1
            continue
        if orderCancelled == True:
            ws.cell(row=eRow, column=8).value = 'Cancelled'
            eRow += 1
            while True:
                try:
                    wb.save(jStatFileName)
                    break
                except PermissionError:
                    exceptPermissionError()
            continue
        if aNF == True:
            ws.cell(row=eRow, column=8).value = 'Appointment not found'
            skipOrder = True
        if skipOrder == True:
            eRow += 1
            continue
        waitLoad2()
        if skipOrder == True:
            ieRefresh()
            try:
                type2(ws.cell(row=eRow, column=7).value)
            except TypeError:
                textmyself.textmyself('Blank ARN - Break')
                break
            waitLoad2()
            if skipOrder == True:
                ws.cell(row=eRow, column=8).value = 'did not load'
                eRow += 1
                continue
        highlightOrderNumber()
        copy()
        ws.cell(row=eRow, column=10).value = pyperclip.paste()
        eRow += 1
        while True:
            try:
                wb.save(jStatFileName)
                break
            except PermissionError:
                exceptPermissionError()
        
    
    
def updateOrderNumber():
    apptManSetUp()
    excelLoop1()

    
def checkWfaClear1():
    if pixelMatch(checkWfa1[0],checkWfa1[1],checkWfa1[2],checkWfa1[3],checkWfa1[4]) or pixelMatch(checkWfa2[0],checkWfa2[1],checkWfa2[2],checkWfa2[3],checkWfa2[4]) \
       or pixelMatch(checkWfa3[0],checkWfa3[1],checkWfa3[2],checkWfa3[3],checkWfa3[4]):
            time.sleep(2)
            pyautogui.press('f8')
            time.sleep(2)
            pyautogui.press('f8')
            if pixelMatch(checkWfa1[0],checkWfa1[1],checkWfa1[2],checkWfa1[3],checkWfa1[4]) or pixelMatch(checkWfa2[0],checkWfa2[1],checkWfa2[2],checkWfa2[3],checkWfa2[4]) \
              or pixelMatch(checkWfa3[0],checkWfa3[1],checkWfa3[2],checkWfa3[3],checkWfa3[4]):
                textmyself.textmyself('Clear WFA')
                pyautogui.confirm('Clear WFA!!')

         
def typeCenter():
    click(wfaCenter[0], wfaCenter[1])
    pyautogui.press('tab')
    pyautogui.typewrite(center)


def copyJobStat():
    pyautogui.mouseDown(jStatDown[0], jStatDown[1])
    pyautogui.mouseUp(jStatUp[0], jStatUp[1])
    copy()
    

def highlightWfaCenter():
    pyautogui.mouseDown(centerDown[0],centerDown[1])
    pyautogui.mouseUp(centerUp[0],centerUp[1])


def fixWfaPrompt():
    fixWfa = ''
    while fixWfa == '':
        fixWfa = pyautogui.confirm(text='WFA is not in the correct Location. Make sure it is the second icon on the taskbar.', title='Fix WFA', buttons=['Fixed', 'Exit'])
        if fixWfa == 'Fixed':
            pyautogui.click(wfaIcon[0],wfaIcon[1]) #wfa taskbar
            pyautogui.hotkey('win', 'up')
            click(ISWR[0],ISWR[1]) # ISWR
            checkCenter()
        if fixWfa == 'Exit':
            fixWfa = pyautogui.confirm(text='Are you sure you want to exit? This will exit out of the program completely.', title='Fix WFA', buttons=['Go Back', 'Exit'])
            if fixWfa == 'Go Back':
                fixWfa = ''
            if fixWfa == 'Exit':
                print('User exited at fix WFA prompt screen.')
                sys.exit()
        

def checkCenter():
    highlightWfaCenter()
    copy()
    wb = load_workbook(filename = jStatFileName)
    wb.get_sheet_names()
    ws = wb.get_sheet_by_name('Report Data')
    ws.cell(row=1, column =20).value = pyperclip.paste()
    while True:
        try:
            wb.save(jStatFileName)
            break
        except PermissionError:
            exceptPermissionError2()
    if ws.cell(row=1, column =20).value != 'CENTER':
        fixWfaPrompt()

        
def updateJobStat():
    wb = load_workbook(filename = jStatFileName)
    wb.get_sheet_names()
    ws = wb.get_sheet_by_name('Report Data')
    wfaTaskbar()
    pyautogui.hotkey('win', 'up')
    click(253, 101) #ISWR
    checkCenter()
    global eRow2
    for i in range(ws.max_row-2):
        pyautogui.press('f8')
        checkWfaClear1()
        try:
            if len(ws.cell(row=eRow2, column=10).value) == 9:
                typeCenter()
                pyautogui.press('tab')
                pyautogui.press('tab')
                type2(ws.cell(row=eRow2, column=10).value)
                time.sleep(.3)
                pyautogui.press('f1')
                time.sleep(1)
                copyJobStat()
                ws.cell(row=eRow2, column =8).value = pyperclip.paste()
                eRow2 += 1
            elif len(ws.cell(row=eRow2, column=10).value) == 8:
                ws.cell(row=eRow2, column =8).value = 'ION'
                eRow2 += 1
        except TypeError:
            ws.cell(row=eRow2, column =8).value = 'BON'
            eRow2 += 1
        while True:
            try:
                wb.save(jStatFileName)
                break
            except PermissionError:
                exceptPermissionError2()


class orderDetailsGUI(tkinter.Tk):
    def __init__(self,parent):
        tkinter.Tk.__init__(self,parent)
        self.parent = parent
        self.initialize()
        

    def initialize(self):
        self.grid()
        menubar = Menu(self)

        fileMenu = Menu(menubar, tearoff=0)
        fileMenu.add_command(label="Select Order Details Report .xlsx File", command=self.onOpen)
        fileMenu.add_command(label="Appointment Scheduler", command=self.openIEappt)
        fileMenu.add_command(label="Websop", command=self.openIEwebsop)
        fileMenu.add_separator()
        fileMenu.add_command(label="Exit", command=self.onQuit)
        menubar.add_cascade(label="File", menu=fileMenu)

        runMenu = Menu(menubar, tearoff=0)
        runMenu.add_command(label="Update Order Number & Job Status", command=self.onBoth)
        runMenu.add_command(label="Just Update Order Number", command=self.onOrder)
        runMenu.add_command(label="Just Update Job Status", command=self.onJobStat)
        menubar.add_cascade(label="Run", menu=runMenu)

        resolutionMenu = Menu(menubar, tearoff=0)
        resolutionMenu.add_command(label="1366x768",command=self.setRes1366)
        resolutionMenu.add_command(label="1680x1050", command=self.setRes1680)
        menubar.add_cascade(label='Resolution', menu=resolutionMenu)

        helpMenu = Menu(menubar, tearoff=0)
        helpMenu.add_command(label="About", command=self.aboutMessage)
        menubar.add_cascade(label="Help", menu=helpMenu)

        self.centerText = tkinter.StringVar()
        self.centerEntry = tkinter.Entry(self,textvariable=self.selectCenter)
        self.centerEntry.grid(column=0,row=0,sticky='EW')
        self.centerEntry.bind("<Return>", self.selectCenter)
        self.centerText.set(u"Type center name here")

        changeCenterButton = tkinter.Button(self,text=u"Change Center", command=self.changeCenterClick)
        changeCenterButton.grid(column=1,row=0,sticky='EW')

        """self.centerVariable = tkinter.StringVar()
        self.entry = tkinter.Entry(self,textvariable=self.centerVariable)
        self.entry.grid(column=0,row=0,sticky='EW')
        self.entry.bind("<Return>", self.onPressEnter)
        self.centerVariable.set(u"Enter center here.")"""

        self.rowVariable = tkinter.StringVar()
        self.rowEntry = tkinter.Entry(self,textvariable=self.rowVariable)
        self.rowEntry.grid(column=3,row=0,sticky='EW')
        self.rowEntry.bind("<Return>", self.onRowEnter)
        self.rowVariable.set(u"Select start row")

        """self.rowVariable = Spinbox(self, from_=3, to=1000, command =self.onRowEnter)
        self.entry.bind("<Return>", self.onRowEnter)
        self.rowVariable.grid(column=3,row=0)"""

        self.rowStart = tkinter.StringVar()
        rowLabel = tkinter.Label(self,textvariable=self.rowStart,anchor="w",fg="black")
        rowLabel.grid(column=3,row=1,sticky='EW')
        self.rowStart.set(u"row = 3")

        """enterButton = tkinter.Button(self,text=u"Enter", command=self.onEnterClick)
        enterButton.grid(column=1,row=0,stick='EW')"""
        
        self.centerName = tkinter.StringVar()
        centerLabel =tkinter.Label(self,textvariable=self.centerName,anchor="w",fg="black")
        centerLabel.grid(column=0,row=1,columnspan=1,sticky='EW')
        self.centerName.set(u"TYPE CENTER NAME!")

        self.fileName = tkinter.StringVar()
        fileLabel =tkinter.Label(self,textvariable=self.fileName,anchor="w",fg="black",bg="red")
        fileLabel.grid(column=2,row=1,columnspan=1,sticky='EW')
        self.fileName.set(u"SELECT A FILE!")

        self.grid_columnconfigure(0,weight=1)
        self.config(menu=menubar)
        self.resizable(True,False)
        self.update()
        self.geometry(self.geometry())
        self.centerEntry.focus_set()
        self.centerEntry.selection_range(0, tkinter.END)

    def selectCenter(self,event):
        global center
        center = self.centerEntry.get()
        self.centerName.set(u"Center: " + self.centerEntry.get())
        self.centerEntry.focus_set()
        self.centerEntry.selection_range(0, tkinter.END)
        
    def changeCenterClick(self):
        global center
        center = self.centerEntry.get()
        self.centerName.set(u"Center: " + self.centerEntry.get())
        self.centerEntry.focus_set()
        self.centerEntry.selection_range(0, tkinter.END)
                    
    """def onEnterClick(self):
        global center
        self.centerName.set(self.centerVariable.get())
        center = self.centerVariable.get()
        print(center)
        self.entry.focus_set()
        self.entry.selection_range(0, tkinter.END)

    def onPressEnter(self,event):
        global center
        self.centerName.set(self.centerVariable.get())
        center = self.centerVariable.get()
        print(center)
        self.entry.focus_set()
        self.entry.selection_range(0, tkinter.END)"""

    def onOrder(self):
        sleep(1)
        updateOrderNumber()
        textmyself.textmyself('Updating order number for ' + name[44:49] + ' has been completed, Michael.')
        sys.exit()

    def onJobStat(self):
        sleep(1)
        updateJobStat()
        textmyself.textmyself('Updating job stat for ' + name[44:49] + ' has been completed, Michael.')
        sys.exit()

    def onRowEnter(self,event):
        global eRow
        global eRow2
        eRow = int(self.rowVariable.get())
        eRow2 = int(self.rowVariable.get())
        self.rowStart.set(u"row = " + str(eRow))

    def onOpen(self):
        global name
        openReportFile()
        self.fileName.set(name[44:])
        fileLabel =tkinter.Label(self,textvariable=self.fileName,anchor="w",fg="white",bg="blue")
        self.update()

    def onBoth(self):
        sleep(1)
        updateOrderNumber()
        updateJobStat()
        textmyself.textmyself('The order details report for ' + name[44:49] + ' has been completed, Michael.')
        sys.exit()

    def openIEappt(self):
        os.startfile('https://asapp.foss.qintra.com/asapp/faces/index.jsp')

    def openIEwebsop(self):
        os.startfile('https://websop.foss.corp.intranet/websop/jsp/login.jsp')

    def setRes1680(self):
        global resolution
        resolution = 1680
        setResolution()
        print(waitLoad11)
        print(waitLoad22)
        print(waitLoad3)
        print(waitLoad4)

    def setRes1366(self):
        resolution = 1366
        return resolution
        setResolution()

    def aboutMessage(self):
        print('about message dialog/text will go here')

    def onQuit(self):
        self.destroy()
        sys.exit()

                            
if __name__ == "__main__":
    app = orderDetailsGUI(None)
    app.title('Automated Order Details Report')
    app.mainloop()

